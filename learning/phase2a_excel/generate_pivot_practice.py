import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import random
from datetime import date, timedelta

random.seed(42)

# --- Data definition ---
outlets = ["Dubai Mall", "Marina Mall", "Online"]
categories = {
    "Apparel":      [("Linen Shirt", 180, 72), ("Abaya", 320, 128), ("Casual Dress", 210, 84)],
    "Accessories":  [("Leather Bag", 450, 180), ("Sunglasses", 290, 116), ("Silk Scarf", 160, 64)],
    "Footwear":     [("Sneakers", 380, 152), ("Sandals", 220, 88), ("Heels", 340, 136)],
}

# Generate ~108 rows: 6 months x 3 outlets x 3 categories x 2 transactions each
rows = []
start = date(2025, 1, 1)

for month_offset in range(6):           # Jan–Jun 2025
    month_start = date(2025, month_offset + 1, 1)
    for outlet in outlets:
        for category, products in categories.items():
            for _ in range(2):          # 2 transactions per outlet/category/month
                product, price, cogs_unit = random.choice(products)
                # Seasonality: Ramadan (Mar) +20%, summer (Jun) -15%
                if month_offset == 2:
                    units = random.randint(14, 22)
                elif month_offset == 5:
                    units = random.randint(5, 12)
                else:
                    units = random.randint(8, 18)
                # Online outlet tends to have higher volume
                if outlet == "Online":
                    units = int(units * 1.3)
                day = random.randint(1, 28)
                txn_date = date(2025, month_offset + 1, day)
                revenue = round(units * price, 2)
                cogs = round(units * cogs_unit, 2)
                rows.append([txn_date, outlet, category, product, units, price, revenue, cogs])

random.shuffle(rows)

# --- Build workbook ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales Data"

headers = ["Date", "Outlet", "Category", "Product", "Units_Sold", "Unit_Price", "Revenue", "COGS"]
header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write headers
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data
alt_fill = PatternFill("solid", fgColor="F2F6FC")
for row_idx, row in enumerate(rows, 2):
    fill = alt_fill if row_idx % 2 == 0 else None
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = cell_border
        if fill:
            cell.fill = fill
        if col_idx == 1:                # Date column
            cell.number_format = "DD-MMM-YYYY"
            cell.alignment = Alignment(horizontal="center")
        elif col_idx in (5,):           # Units
            cell.alignment = Alignment(horizontal="center")
        elif col_idx in (6, 7, 8):     # Price, Revenue, COGS
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

# Format as Excel Table
last_row = len(rows) + 1
last_col = get_column_letter(len(headers))
table_ref = f"A1:{last_col}{last_row}"
tbl = Table(displayName="SalesData", ref=table_ref)
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(tbl)

# Column widths
widths = [14, 14, 14, 18, 12, 12, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

wb.save("/Users/ashnad/my_career/learning/phase2a_excel/05_pivottable_practice.xlsx")
print(f"Created 05_pivottable_practice.xlsx — {len(rows)} rows of UAE retail transaction data")
