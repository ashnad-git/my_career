"""
Generate 04_index_match_practice.xlsx — INDEX-MATCH + IFERROR practice file.
Run: python3 learning/phase2a_excel/generate_index_match_practice.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE_DARK  = "1e3a5f"
BLUE_LIGHT = "d6e4f0"
GREY_ROW   = "f2f2f2"
YELLOW     = "fff2cc"


def header_style():
    return {
        "font":      Font(bold=True, color="FFFFFF", size=10),
        "fill":      PatternFill("solid", fgColor=BLUE_DARK),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Employee data ──────────────────────────────────────────────────────────────
# Columns: ID | Name | Department | Role | Salary (AED) | Location
# Deliberately shuffled — not sorted by ID — so MATCH must actually search

employees = [
    ("E-014", "Hana Al Rashid",       "Treasury",       "Analyst",        13000,  "Dubai"),
    ("E-022", "Ayesha Al Ketbi",      "Accounting",     "Senior Analyst", 16000,  "Dubai"),
    ("E-007", "Rajesh Kumar",         "Finance",        "Analyst",        11000,  "Sharjah"),
    ("E-019", "Sanjay Mehta",         "Finance",        "Analyst",        10500,  "Dubai"),
    ("E-026", "Hessa Al Suwaidi",     "FP&A",           "Analyst",        12500,  "Dubai"),
    ("E-011", "Khalid Al Suwaidi",    "FP&A",           "Manager",        25000,  "Dubai"),
    ("E-003", "Omar Hassan Al Farsi", "Accounting",     "Analyst",        12000,  "Dubai"),
    ("E-028", "Shaikha Al Muhairi",   "Finance",        "Director",       45000,  "Abu Dhabi"),
    ("E-016", "Layla Al Hajri",       "FP&A",           "Senior Analyst", 19000,  "Abu Dhabi"),
    ("E-030", "Zainab Al Darmaki",    "Treasury",       "Analyst",        12000,  "Sharjah"),
    ("E-005", "James Mitchell",       "Internal Audit", "Senior Manager", 28000,  "Abu Dhabi"),
    ("E-024", "Noura Al Shamsi",      "Finance",        "Manager",        23500,  "Abu Dhabi"),
    ("E-008", "Aisha Al Hamdan",      "Accounting",     "Senior Analyst", 15500,  "Dubai"),
    ("E-017", "Daniel Park",          "Internal Audit", "Manager",        23000,  "Dubai"),
    ("E-001", "Ahmed Al Rashidi",     "FP&A",           "Senior Analyst", 18000,  "Dubai"),
    ("E-025", "Mark Williams",        "Treasury",       "Manager",        26000,  "Dubai"),
    ("E-012", "Nadia El Sayed",       "Finance",        "Senior Analyst", 17000,  "Dubai"),
    ("E-006", "Sara Al Mansoori",     "FP&A",           "Analyst",        13500,  "Dubai"),
    ("E-029", "Lucas Martin",         "Internal Audit", "Analyst",        11000,  "Dubai"),
    ("E-020", "Reem Al Nuaimi",       "Treasury",       "Senior Analyst", 15000,  "Abu Dhabi"),
    ("E-009", "Mohammed Al Marzouqi","Treasury",        "Manager",        24000,  "Abu Dhabi"),
    ("E-027", "Anand Iyer",           "Accounting",     "Manager",        22000,  "Dubai"),
    ("E-004", "Priya Nair",           "Treasury",       "Senior Analyst", 16000,  "Dubai"),
    ("E-023", "Arjun Pillai",         "Internal Audit", "Senior Analyst", 17500,  "Dubai"),
    ("E-013", "Thomas Clarke",        "Accounting",     "Manager",        21000,  "Abu Dhabi"),
    ("E-021", "Patrick O'Brien",      "FP&A",           "Director",       42000,  "Dubai"),
    ("E-002", "Fatima Al Zaabi",      "Finance",        "Manager",        22000,  "Abu Dhabi"),
    ("E-018", "Mariam Al Blooshi",    "Accounting",     "Analyst",        11500,  "Sharjah"),
    ("E-010", "Deepa Krishnan",       "Internal Audit", "Analyst",        12500,  "Dubai"),
    ("E-015", "Vikram Sharma",        "Finance",        "Senior Manager", 30000,  "Dubai"),
]

wb = openpyxl.Workbook()

# ── Sheet 1: DATA ──────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Data"

ws["A1"] = "INDEX-MATCH + IFERROR Practice — Zara & Co. Finance Team (Fictional)"
ws["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
ws.merge_cells("A1:F1")

ws["A2"] = "30 employees across Finance, FP&A, Treasury, Accounting, Internal Audit | Salaries in AED"
ws["A2"].font = Font(italic=True, size=9, color="666666")
ws.merge_cells("A2:F2")

ws["A3"] = "NOTE: Data is intentionally NOT sorted by Employee ID — MATCH must actually search."
ws["A3"].font = Font(italic=True, size=9, color="AA0000")
ws.merge_cells("A3:F3")

headers = ["Employee_ID", "Name", "Department", "Role", "Salary (AED)", "Location"]
col_widths = [14, 24, 18, 18, 14, 14]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=5, column=i, value=h)
    apply_style(cell, **header_style())
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[5].height = 22

for r, row in enumerate(employees, 6):
    fill = PatternFill("solid", fgColor=GREY_ROW) if r % 2 == 0 else None
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if c != 2 else "left")
        if fill:
            cell.fill = fill
        if c == 5:
            cell.number_format = "#,##0"

ws.freeze_panes = "A6"

# ── Sheet 2: EXERCISES ─────────────────────────────────────────────────────────
wx = wb.create_sheet("Exercises")

wx["A1"] = "INDEX-MATCH + IFERROR Exercises"
wx["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
wx.merge_cells("A1:D1")

wx["A2"] = "Data is on the 'Data' sheet (rows 6–35). Write your formula in column D."
wx["A2"].font = Font(italic=True, size=9, color="666666")
wx.merge_cells("A2:D2")

# Lookup reference box — so students don't have to memorise column positions
wx["F1"] = "Column Reference (Data sheet)"
wx["F1"].font = Font(bold=True, size=9, color=BLUE_DARK)
wx.merge_cells("F1:G1")

ref_rows = [
    ("A (col 1)", "Employee_ID"),
    ("B (col 2)", "Name"),
    ("C (col 3)", "Department"),
    ("D (col 4)", "Role"),
    ("E (col 5)", "Salary (AED)"),
    ("F (col 6)", "Location"),
]
for i, (col, name) in enumerate(ref_rows, 2):
    wx.cell(row=i, column=6, value=col).font = Font(size=9, bold=True)
    wx.cell(row=i, column=7, value=name).font = Font(size=9)

wx.column_dimensions["F"].width = 14
wx.column_dimensions["G"].width = 18

ex_headers = ["#", "Question", "Difficulty", "Your Formula / Answer"]
ex_widths  = [4,   68,         14,           40]
for i, (h, w) in enumerate(zip(ex_headers, ex_widths), 1):
    cell = wx.cell(row=4, column=i, value=h)
    apply_style(cell, **header_style())
    cell.border = thin_border()
    wx.column_dimensions[get_column_letter(i)].width = w

exercises = [
    # Level 1 — right lookup (same direction as VLOOKUP, but using INDEX-MATCH)
    (1,  "Employee ID E-008 is in cell H2. Return their full Name.\n"
         "(This is a right lookup — same as VLOOKUP, but use INDEX-MATCH.)",
         "Level 1"),
    (2,  "Return the Department for Employee ID E-015.",
         "Level 1"),
    (3,  "Return the Salary (AED) for Employee ID E-022.",
         "Level 1"),

    # Level 2 — left lookup (VLOOKUP cannot do this)
    (4,  "You know the name is 'Omar Hassan Al Farsi'. Return their Employee ID.\n"
         "Employee ID is in column A — to the LEFT of the Name column.\n"
         "VLOOKUP cannot do this. INDEX-MATCH can.",
         "Level 2 ★"),
    (5,  "Given the name 'Priya Nair', return their Role.",
         "Level 2 ★"),
    (6,  "Given the name 'James Mitchell', return their Location.",
         "Level 2 ★"),

    # Level 3 — IFERROR
    (7,  "Employee ID 'E-099' does not exist in the table.\n"
         "Write INDEX-MATCH wrapped in IFERROR to return 'Employee not found' instead of #N/A.",
         "Level 3"),
    (8,  "Return the Salary for 'Khalid Al Suwaidi'. Then wrap in IFERROR so that\n"
         "if the name is mistyped and not found, the cell shows 0 instead of an error.",
         "Level 3"),

    # Level 4 — open-ended / combined
    (9,  "Find the NAME of the highest-paid employee in the entire company.\n"
         "Hint: combine INDEX-MATCH with MAX().",
         "Level 4 ★★"),
    (10, "Find the NAME of the highest-paid employee in the FP&A department only.\n"
         "Hint: you cannot use MAX directly — think about which formula filters by department.",
         "Level 4 ★★"),
    (11, "How many employees are in each department? Use COUNTIF for each of the 5 departments.\n"
         "Then use INDEX-MATCH to look up the count for any department name you type in a cell.",
         "Level 4 ★★"),
]

for r, (num, q, lvl) in enumerate(exercises, 5):
    fill = PatternFill("solid", fgColor=GREY_ROW) if r % 2 == 0 else None
    for c, val in enumerate([num, q, lvl, ""], 1):
        cell = wx.cell(row=r, column=c, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=(c == 2), horizontal="center" if c != 2 else "left")
        if fill:
            cell.fill = fill
    wx.row_dimensions[r].height = 48

wx.freeze_panes = "A5"

# ── Sheet 3: ANSWER KEY (hidden) ───────────────────────────────────────────────
wa = wb.create_sheet("Answer Key")
wa.sheet_state = "hidden"

wa["A1"] = "Answer Key — INDEX-MATCH + IFERROR Practice"
wa["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
wa.merge_cells("A1:C1")

ans_headers = ["#", "Answer", "Formula"]
ans_widths  = [4,   28,       90]
for i, (h, w) in enumerate(zip(ans_headers, ans_widths), 1):
    cell = wa.cell(row=3, column=i, value=h)
    apply_style(cell, **header_style())
    wa.column_dimensions[get_column_letter(i)].width = w

answers = [
    (1,  "Aisha Al Hamdan",
         '=INDEX(Data!B6:B35, MATCH("E-008", Data!A6:A35, 0))'),
    (2,  "Finance",
         '=INDEX(Data!C6:C35, MATCH("E-015", Data!A6:A35, 0))'),
    (3,  "16,000",
         '=INDEX(Data!E6:E35, MATCH("E-022", Data!A6:A35, 0))'),
    (4,  "E-003",
         '=INDEX(Data!A6:A35, MATCH("Omar Hassan Al Farsi", Data!B6:B35, 0))'),
    (5,  "Senior Analyst",
         '=INDEX(Data!D6:D35, MATCH("Priya Nair", Data!B6:B35, 0))'),
    (6,  "Abu Dhabi",
         '=INDEX(Data!F6:F35, MATCH("James Mitchell", Data!B6:B35, 0))'),
    (7,  "Employee not found",
         '=IFERROR(INDEX(Data!B6:B35, MATCH("E-099", Data!A6:A35, 0)), "Employee not found")'),
    (8,  "25,000 (or 0 if name wrong)",
         '=IFERROR(INDEX(Data!E6:E35, MATCH("Khalid Al Suwaidi", Data!B6:B35, 0)), 0)'),
    (9,  "Shaikha Al Muhairi (45,000 AED)",
         '=INDEX(Data!B6:B35, MATCH(MAX(Data!E6:E35), Data!E6:E35, 0))'),
    (10, "Patrick O\'Brien (42,000 AED)",
         '=INDEX(Data!B6:B35, MATCH(MAXIFS(Data!E6:E35, Data!C6:C35, "FP&A"), Data!E6:E35, 0))'),
    (11, "Finance=7, FP&A=6, Treasury=6, Accounting=6, Internal Audit=5",
         '=COUNTIF(Data!C6:C35,"Finance") etc. Then INDEX-MATCH over a summary table.'),
]

for r, (num, ans, formula) in enumerate(answers, 4):
    wa.cell(row=r, column=1, value=num)
    wa.cell(row=r, column=2, value=ans)
    cell = wa.cell(row=r, column=3, value=formula)
    cell.font = Font(name="Courier New", size=9)
    for c in range(1, 4):
        wa.cell(row=r, column=c).border = thin_border()
        wa.row_dimensions[r].height = 20

wb.active = wb["Data"]

path = "/Users/ashnad/my_career/learning/phase2a_excel/04_index_match_practice.xlsx"
wb.save(path)
print(f"Saved: {path}")
print(f"30 employees | 11 exercises | Answer Key hidden")
