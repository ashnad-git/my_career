"""
Generate 07_conditional_formatting_practice.xlsx
Run: python3 learning/phase2a_excel/generate_cf_practice.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

BLUE_DARK = "1e3a5f"
GREY_ROW  = "f2f2f2"
YELLOW    = "fff2cc"


def header_style():
    return {
        "font":      Font(bold=True, color="FFFFFF", size=10),
        "fill":      PatternFill("solid", fgColor=BLUE_DARK),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


wb = openpyxl.Workbook()

# ── Sheet 1: VARIANCE TABLE ────────────────────────────────────────────────────
ws = wb.active
ws.title = "Variance Table"

ws["A1"] = "Conditional Formatting Practice — Zara & Co. P&L Variance (Jan 2025)"
ws["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
ws.merge_cells("A1:E1")

ws["A2"] = "Variance = Actual − Budget  |  For cost lines: positive variance = over budget (unfavourable)"
ws["A2"].font = Font(italic=True, size=9, color="666666")
ws.merge_cells("A2:E2")

headers = ["Line Item", "Budget (AED)", "Actual (AED)", "Variance (AED)", "Variance %"]
col_widths = [22, 16, 16, 16, 14]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=4, column=i, value=h)
    apply_style(cell, **header_style())
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[4].height = 22

# Data: Line Item | Budget | Actual
# Variance and Variance % are written as formulas
rows = [
    ("Revenue",   500000,  480000),
    ("COGS",      200000,  225000),
    ("Gross Profit", 300000, 255000),
    ("Salaries",  150000,  148000),
    ("Marketing",  50000,   71000),
    ("Rent",       80000,   80000),
    ("Other OpEx", 20000,   31000),
    ("EBITDA",     20000,  -44000),
]

for r, (label, budget, actual) in enumerate(rows, 5):
    fill = PatternFill("solid", fgColor=GREY_ROW) if r % 2 == 0 else None

    # Line Item
    c1 = ws.cell(row=r, column=1, value=label)
    c1.font = Font(bold=(label in ("Gross Profit", "EBITDA")))
    c1.border = thin_border()
    c1.alignment = Alignment(horizontal="left")
    if fill: c1.fill = fill

    # Budget
    c2 = ws.cell(row=r, column=2, value=budget)
    c2.number_format = "#,##0"
    c2.border = thin_border()
    c2.alignment = Alignment(horizontal="right")
    if fill: c2.fill = fill

    # Actual
    c3 = ws.cell(row=r, column=3, value=actual)
    c3.number_format = "#,##0"
    c3.border = thin_border()
    c3.alignment = Alignment(horizontal="right")
    if fill: c3.fill = fill

    # Variance $ — formula: Actual - Budget
    c4 = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    c4.number_format = "#,##0"
    c4.border = thin_border()
    c4.alignment = Alignment(horizontal="right")
    if fill: c4.fill = fill

    # Variance % — formula: Variance / Budget
    c5 = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},0)")
    c5.number_format = "0.0%"
    c5.border = thin_border()
    c5.alignment = Alignment(horizontal="right")
    if fill: c5.fill = fill

# Totals label
ws.row_dimensions[13].height = 8  # small spacer

ws.freeze_panes = "A5"

# ── Sheet 2: TASKS ─────────────────────────────────────────────────────────────
wt = wb.create_sheet("Tasks")

wt["A1"] = "Conditional Formatting — Your Tasks"
wt["A1"].font = Font(bold=True, size=13, color=BLUE_DARK)
wt.merge_cells("A1:B1")
wt.column_dimensions["A"].width = 12
wt.column_dimensions["B"].width = 72

tasks = [
    ("TASK 1",
     "HIGHLIGHT CELL RULES — Variance (AED) column (D5:D12)\n"
     "Rule 1: If value < 0  →  Red fill (#FFC7CE), dark red text (#9C0006)\n"
     "Rule 2: If value > 0  →  Green fill (#C6EFCE), dark green text (#276221)\n"
     "Rule 3: If value = 0  →  Yellow fill (#FFEB9C), dark yellow text (#9C6500)\n\n"
     "How: Select D5:D12 → Home → Conditional Formatting → Highlight Cells Rules → Less Than / Greater Than / Equal To"),

    ("TASK 2",
     "DATA BARS — Budget (AED) column (B5:B12)\n"
     "Apply a blue data bar so you can visually compare budget sizes at a glance.\n\n"
     "How: Select B5:B12 → Home → Conditional Formatting → Data Bars → Blue Data Bar"),

    ("TASK 3",
     "ICON SET — Variance % column (E5:E12)\n"
     "Apply a 3-icon traffic light set:\n"
     "  Green circle  = Variance % >= 0  (on/above budget for revenue, under for costs)\n"
     "  Yellow circle = Variance % between -5% and 0\n"
     "  Red circle    = Variance % < -5%\n\n"
     "How: Select E5:E12 → Conditional Formatting → Icon Sets → 3 Traffic Lights\n"
     "Then: Manage Rules → Edit → set thresholds manually (type = Number, values 0 and -0.05)"),

    ("TASK 4 ★",
     "CUSTOM FORMULA RULE — Entire row (A5:E12)\n"
     "If the Variance % in column E is worse than -10%, highlight the ENTIRE ROW in light red.\n"
     "This requires a custom formula so you can check column E while coloring columns A–E.\n\n"
     "How: Select A5:E12 → Conditional Formatting → New Rule → 'Use a formula'\n"
     "Formula: =$E5<-0.10\n"
     "Format: Fill → light red (#FFD7D7)\n\n"
     "Why the $ before E but not before 5?\n"
     "  $E locks the column (always check column E, not whatever column we're currently in)\n"
     "  No $ on 5 lets the row number move as the rule checks each row"),
]

row = 3
for tag, desc in tasks:
    # Tag cell
    tag_cell = wt.cell(row=row, column=1, value=tag)
    tag_cell.font = Font(bold=True, color="FFFFFF", size=10)
    tag_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    tag_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    tag_cell.border = thin_border()

    # Description cell
    desc_cell = wt.cell(row=row, column=2, value=desc)
    desc_cell.alignment = Alignment(vertical="top", wrap_text=True)
    desc_cell.border = thin_border()
    desc_cell.fill = PatternFill("solid", fgColor="F7FBFF")

    wt.row_dimensions[row].height = 100
    row += 2  # gap between tasks

# Key concept box
wt.cell(row=row + 1, column=1, value="KEY CONCEPT").font = Font(bold=True, color=BLUE_DARK)
wt.merge_cells(f"A{row+1}:B{row+1}")
concept = (
    "The $ anchor in custom formula rules:\n"
    "  =$E5<-0.10   ← correct: locks column E, row moves per row\n"
    "  =E5<-0.10    ← wrong: both move — formula checks wrong column as it scans right\n"
    "  =$E$5<-0.10  ← wrong: nothing moves — every row checks the same cell (row 5)\n\n"
    "Rule: lock the column ($E), never lock the row, when applying to a multi-column range."
)
concept_cell = wt.cell(row=row + 2, column=1, value=concept)
concept_cell.alignment = Alignment(wrap_text=True, vertical="top")
concept_cell.fill = PatternFill("solid", fgColor=YELLOW)
concept_cell.font = Font(size=9)
wt.merge_cells(f"A{row+2}:B{row+2}")
wt.row_dimensions[row + 2].height = 100

wb.active = wb["Variance Table"]

path = "/Users/ashnad/my_career/learning/phase2a_excel/07_conditional_formatting_practice.xlsx"
wb.save(path)
print(f"Saved: {path}")
print("Sheets: Variance Table (data + formulas) | Tasks (4 CF exercises)")
