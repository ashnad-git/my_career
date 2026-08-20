"""
Generate Phase 2A Excel practice files.
Run: python3 learning/phase2a_excel/generate_practice_files.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter

# ─── Colour palette ────────────────────────────────────────────────────────────
BLUE_DARK  = "1e3a5f"   # headers
BLUE_LIGHT = "d6e4f0"   # header fill
GREY_ROW   = "f2f2f2"   # alternating row fill
GREEN      = "c6efce"   # correct answer highlight
YELLOW     = "fff2cc"   # instruction / note


def header_style():
    return {
        "font":      Font(bold=True, color="FFFFFF", size=10),
        "fill":      PatternFill("solid", fgColor=BLUE_DARK),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ──────────────────────────────────────────────────────────────────────────────
# FILE 01 — SUMIFS  (medium difficulty: 30 rows, 3 criteria fields)
# ──────────────────────────────────────────────────────────────────────────────

def build_sumifs_file():
    wb = openpyxl.Workbook()

    # ── Sheet 1: DATA ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "SUMIFS Practice — Zara & Co. UAE (Fictional)"
    ws["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
    ws.merge_cells("A1:F1")
    ws["A2"] = "Source: Monthly departmental spend data, Jan–Jun 2025 (AED)"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("A2:F2")

    headers = ["Month", "Quarter", "Region", "Department", "Category", "Amount (AED)"]
    col_widths = [12, 10, 14, 16, 18, 16]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=i, value=h)
        apply_style(cell, **header_style())
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 22

    # 30 rows of realistic UAE retail data
    data = [
        # Month   Qtr    Region         Dept          Category       Amount
        ("Jan",  "Q1", "Dubai",       "Sales",      "Salaries",    42000),
        ("Jan",  "Q1", "Dubai",       "Sales",      "Commission",   8500),
        ("Jan",  "Q1", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("Jan",  "Q1", "Abu Dhabi",   "Marketing",  "Advertising",  15000),
        ("Jan",  "Q1", "Dubai",       "Finance",    "Salaries",    28000),
        ("Feb",  "Q1", "Dubai",       "Sales",      "Salaries",    42000),
        ("Feb",  "Q1", "Dubai",       "Sales",      "Commission",  12000),
        ("Feb",  "Q1", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("Feb",  "Q1", "Abu Dhabi",   "Marketing",  "Advertising",  18000),
        ("Feb",  "Q1", "Dubai",       "Finance",    "Salaries",    28000),
        ("Mar",  "Q1", "Dubai",       "Sales",      "Salaries",    42000),
        ("Mar",  "Q1", "Dubai",       "Sales",      "Commission",  21000),
        ("Mar",  "Q1", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("Mar",  "Q1", "Abu Dhabi",   "Marketing",  "Advertising",  22000),
        ("Mar",  "Q1", "Dubai",       "Finance",    "Salaries",    28000),
        ("Apr",  "Q2", "Dubai",       "Sales",      "Salaries",    42000),
        ("Apr",  "Q2", "Dubai",       "Sales",      "Commission",   6000),
        ("Apr",  "Q2", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("Apr",  "Q2", "Abu Dhabi",   "Marketing",  "Advertising",   9000),
        ("Apr",  "Q2", "Dubai",       "Finance",    "Salaries",    28000),
        ("May",  "Q2", "Dubai",       "Sales",      "Salaries",    42000),
        ("May",  "Q2", "Dubai",       "Sales",      "Commission",   5500),
        ("May",  "Q2", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("May",  "Q2", "Abu Dhabi",   "Marketing",  "Advertising",   7500),
        ("May",  "Q2", "Dubai",       "Finance",    "Salaries",    28000),
        ("Jun",  "Q2", "Dubai",       "Sales",      "Salaries",    42000),
        ("Jun",  "Q2", "Dubai",       "Sales",      "Commission",   7000),
        ("Jun",  "Q2", "Abu Dhabi",   "Sales",      "Salaries",    38000),
        ("Jun",  "Q2", "Abu Dhabi",   "Marketing",  "Advertising",  11000),
        ("Jun",  "Q2", "Dubai",       "Finance",    "Salaries",    28000),
    ]

    for r, row in enumerate(data, 5):
        fill = PatternFill("solid", fgColor=GREY_ROW) if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
            if c == 6:  # Amount column
                cell.number_format = "#,##0"

    ws.freeze_panes = "A5"

    # ── Sheet 2: EXERCISES ─────────────────────────────────────────────────────
    wx = wb.create_sheet("Exercises")

    wx["A1"] = "SUMIFS Exercises — type your formula in column D"
    wx["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
    wx.merge_cells("A1:D1")

    wx["A2"] = "Data is on the 'Data' sheet (rows 5–34). All answers in AED."
    wx["A2"].font = Font(italic=True, size=9, color="666666")
    wx.merge_cells("A2:D2")

    ex_headers = ["#", "Question", "Difficulty", "Your Formula / Answer"]
    ex_widths  = [4,   62,          14,            38]
    for i, (h, w) in enumerate(zip(ex_headers, ex_widths), 1):
        cell = wx.cell(row=4, column=i, value=h)
        apply_style(cell, **header_style())
        cell.border = thin_border()
        wx.column_dimensions[get_column_letter(i)].width = w

    exercises = [
        # Level 1 — single criteria (warmup, like your session practice)
        (1,  "Total Commission paid across ALL months",                             "Level 1"),
        (2,  "Total spend in Abu Dhabi across ALL months and departments",          "Level 1"),
        (3,  "Total Salaries paid in Q1 (all regions, all departments)",            "Level 1"),

        # Level 2 — two criteria
        (4,  "Total Commission for Dubai Sales in Feb",                             "Level 2"),
        (5,  "Total Advertising spend in Q2 (all regions)",                        "Level 2"),
        (6,  "Total Finance Salaries in Q1",                                        "Level 2"),
        (7,  "Total Sales spend (all categories) in Abu Dhabi for Mar",            "Level 2"),

        # Level 3 — three criteria
        (8,  "Total Salaries in Dubai Sales department in Q2",                     "Level 3"),
        (9,  "Total Marketing Advertising spend in Abu Dhabi for Q1",              "Level 3"),
        (10, "Total Commission in Dubai in months where commission > 10,000",      "Level 3 ★"),

        # Level 4 — open-ended (no formula hint)
        (11, "Which quarter had higher total spend — Q1 or Q2? Prove it with SUMIFS.", "Level 4 ★★"),
        (12, "What % of total spend is Dubai vs Abu Dhabi? Use SUMIFS + division.", "Level 4 ★★"),
    ]

    for r, (num, q, lvl) in enumerate(exercises, 5):
        fill = PatternFill("solid", fgColor=GREY_ROW) if r % 2 == 0 else None
        for c, val in enumerate([num, q, lvl, ""], 1):
            cell = wx.cell(row=r, column=c, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 2))
            if fill:
                cell.fill = fill
        wx.row_dimensions[r].height = 28

    wx.freeze_panes = "A5"

    # ── Sheet 3: ANSWER KEY (hidden — unhide to check) ─────────────────────────
    wa = wb.create_sheet("Answer Key")
    wb.active = wa
    wa.sheet_state = "hidden"

    wa["A1"] = "Answer Key — SUMIFS Practice"
    wa["A1"].font = Font(bold=True, size=12, color=BLUE_DARK)
    wa.merge_cells("A1:C1")

    ans_headers = ["#", "Answer (AED)", "Formula"]
    ans_widths  = [4,   18,             70]
    for i, (h, w) in enumerate(zip(ans_headers, ans_widths), 1):
        cell = wa.cell(row=3, column=i, value=h)
        apply_style(cell, **header_style())
        wa.column_dimensions[get_column_letter(i)].width = w

    answers = [
        (1,   60000,  '=SUMIFS(Data!F5:F34,Data!E5:E34,"Commission")'),
        (2,  237500,  '=SUMIFS(Data!F5:F34,Data!C5:C34,"Abu Dhabi")'),
        (3,  510000,  '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q1",Data!E5:E34,"Salaries")'),
        (4,   12000,  '=SUMIFS(Data!F5:F34,Data!E5:E34,"Commission",Data!C5:C34,"Dubai",Data!A5:A34,"Feb")'),
        (5,   27500,  '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q2",Data!E5:E34,"Advertising")'),
        (6,   84000,  '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q1",Data!D5:D34,"Finance",Data!E5:E34,"Salaries")'),
        (7,   38000,  '=SUMIFS(Data!F5:F34,Data!C5:C34,"Abu Dhabi",Data!D5:D34,"Sales",Data!A5:A34,"Mar")'),
        (8,  168000,  '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q2",Data!C5:C34,"Dubai",Data!D5:D34,"Sales",Data!E5:E34,"Salaries")'),
        (9,   55000,  '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q1",Data!C5:C34,"Abu Dhabi",Data!D5:D34,"Marketing",Data!E5:E34,"Advertising")'),
        (10,  33000,  '=SUMIFS(Data!F5:F34,Data!C5:C34,"Dubai",Data!E5:E34,"Commission",Data!F5:F34,">10000")'),
        (11,  "Q1=583,500 | Q2=297,000 | Q1 higher", '=SUMIFS(Data!F5:F34,Data!B5:B34,"Q1") and =SUMIFS(Data!F5:F34,Data!B5:B34,"Q2")'),
        (12,  "Dubai≈71% | Abu Dhabi≈29%", '=SUMIFS(Data!F5:F34,Data!C5:C34,"Dubai")/SUM(Data!F5:F34)'),
    ]

    for r, (num, ans, formula) in enumerate(answers, 4):
        wa.cell(row=r, column=1, value=num)
        wa.cell(row=r, column=2, value=ans).number_format = "#,##0" if isinstance(ans, int) else "@"
        wa.cell(row=r, column=3, value=formula).font = Font(name="Courier New", size=9)
        for c in range(1, 4):
            wa.cell(row=r, column=c).border = thin_border()

    # Set Data as active sheet when file opens
    wb.active = wb["Data"]

    path = "/Users/ashnad/my_career/learning/phase2a_excel/01_SUMIFS_practice.xlsx"
    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    build_sumifs_file()
