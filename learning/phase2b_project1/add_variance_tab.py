"""
Builds Variance Analysis tab template — with Fav/Unfav label column.
Run: python3 learning/phase2b_project1/add_variance_tab.py

Source: P&L Summary tab
  Budget cols:  C=Jan, D=Feb, E=Mar
  Actual cols:  P=Jan, Q=Feb, R=Mar
  P&L rows:     6=Revenue, 7=COGS, 8=GP, 9=GM%, 11=OpEx, 12=EBITDA, 13=EBITDA%

Column layout (5 cols per month block):
  A          = Line Item
  B C D E F  = Jan  (Budget | Actual | $ Var | Fav/Unfav | % Var)
  H I J K L  = Feb
  N O P Q R  = Mar
  T U V W X  = Q1 Total
  G, M, S    = narrow spacers

Fav/Unfav logic:
  Revenue, GP, EBITDA, margins → Fav if variance > 0
  COGS, OpEx                   → Fav if variance < 0
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILE = "/Users/ashnad/my_career/learning/phase2b_project1/08_fpa_model.xlsx"

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_GREEN = "E2EFDA"
SECTION_BG  = "D6E4F0"
LABEL_BG    = "F2F2F2"
WHITE       = "FFFFFF"
BUDGET_BG   = "DCE6F1"
ACTUAL_BG   = "FFF2CC"
VAR_BG      = "E2EFDA"
TOTAL_BG    = "203864"


def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def set_cell(ws, row, col, value, bg=None, bold=False, colour="000000",
             size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


wb = openpyxl.load_workbook(FILE)

if "Variance Analysis" in wb.sheetnames:
    tab_idx = wb.sheetnames.index("Variance Analysis")
    del wb["Variance Analysis"]
else:
    tab_idx = 5

ws = wb.create_sheet("Variance Analysis", tab_idx)
ws.sheet_properties.tabColor = "FF0000"

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 22
# 5-col blocks: B-F (Jan), H-L (Feb), N-R (Mar), T-X (Q1)
# spacers: G, M, S
BLOCK_COLS = [2, 8, 14, 20]   # start col of each block (1-indexed)
SPACER_COLS = [7, 13, 19]

for start in BLOCK_COLS:
    for offset in range(5):
        col_letter = get_column_letter(start + offset)
        ws.column_dimensions[col_letter].width = 10 if offset == 3 else 13  # Fav/Unfav narrower
for sc in SPACER_COLS:
    ws.column_dimensions[get_column_letter(sc)].width = 2

# ── Row 1: Title ──────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 24
c = ws.cell(row=1, column=1, value="ZARA & CO. — VARIANCE ANALYSIS FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:X1")

# ── Row 2: Sub-header ─────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 14
c = ws.cell(row=2, column=1,
            value="Actual vs Budget — Q1 YTD. $ Var = Actual − Budget. "
                  "Fav/Unfav: revenue & profit lines Fav if positive; cost lines Fav if negative.")
c.fill = fill(MID_BLUE)
c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A2:X2")

# ── Row 3: Colour key ─────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 14
set_cell(ws, 3, 1, "Colour key:", bold=True, size=9, colour=DARK_BLUE)
for col, label, bg, fg in [
    (2, "  Budget  ", BUDGET_BG, "1F497D"),
    (3, "  Actual  ", ACTUAL_BG, "7F6000"),
    (4, "  Fav (green) / Unfav (red) — conditional formatting applied by Ashnad  ", VAR_BG, "375623"),
]:
    c = ws.cell(row=3, column=col, value=label)
    c.fill = fill(bg); c.font = Font(color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 4: Spacer ─────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 8

# ── Row 5: Month group headers ────────────────────────────────────────────────
ws.row_dimensions[5].height = 16
MONTH_LABELS = [("January", 2, 6), ("February", 8, 12), ("March", 14, 18), ("Q1 Total", 20, 24)]
for label, start, end in MONTH_LABELS:
    bg = TOTAL_BG if "Q1" in label else DARK_BLUE
    c = ws.cell(row=5, column=start, value=label)
    c.fill = fill(bg)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=5, start_column=start, end_row=5, end_column=end)

# ── Row 6: Sub-headers ────────────────────────────────────────────────────────
ws.row_dimensions[6].height = 14
SUB_HDRS = ["Budget", "Actual", "$ Variance", "Fav/Unfav", "% Variance"]
SUB_BG   = [BUDGET_BG, ACTUAL_BG, VAR_BG, VAR_BG, VAR_BG]
SUB_FG   = ["1F497D", "7F6000", "375623", "375623", "375623"]

for blk_start in BLOCK_COLS:
    is_q1 = blk_start == 20
    for i, (hdr, bg, fg) in enumerate(zip(SUB_HDRS, SUB_BG, SUB_FG)):
        c = ws.cell(row=6, column=blk_start+i, value=hdr)
        c.fill = fill(TOTAL_BG if is_q1 else bg)
        c.font = Font(bold=True, color=WHITE if is_q1 else fg, size=8, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

# ── P&L rows ──────────────────────────────────────────────────────────────────
PL_ROWS = [
    (7,  "Revenue",          "#,##0",  False),
    (8,  "Less: COGS",       "#,##0",  False),
    (9,  "Gross Profit",     "#,##0",  True),
    (10, "Gross Margin %",   "0.0%",   True),
    (12, "Less: OpEx",       "#,##0",  False),
    (13, "EBITDA",           "#,##0",  True),
    (14, "EBITDA Margin %",  "0.0%",   True),
]

ws.row_dimensions[11].height = 6   # spacer between GP and OpEx

for tab_row, label, fmt, is_derived in PL_ROWS:
    ws.row_dimensions[tab_row].height = 17
    bold_label = is_derived or label in ["Revenue", "EBITDA"]

    # Col A: label
    set_cell(ws, tab_row, 1, label,
             bg=SECTION_BG if is_derived else LABEL_BG,
             bold=bold_label,
             colour=DARK_BLUE if is_derived else "404040")

    # 4 blocks: Jan, Feb, Mar, Q1
    for blk_start in BLOCK_COLS:
        is_q1 = blk_start == 20

        for offset, (cell_bg, cell_fg) in enumerate([
            (BUDGET_BG, "1F497D"),   # Budget
            (ACTUAL_BG, "7F6000"),   # Actual
            (VAR_BG,    "375623"),   # $ Variance
            (VAR_BG,    "375623"),   # Fav/Unfav
            (VAR_BG,    "375623"),   # % Variance
        ]):
            c = ws.cell(row=tab_row, column=blk_start+offset, value=None)
            c.fill = fill(TOTAL_BG if is_q1 else cell_bg)
            c.font = Font(color=WHITE if is_q1 else cell_fg,
                          size=10, name="Calibri", bold=bold_label if is_q1 else False)
            c.alignment = Alignment(horizontal="center", vertical="center")
            # Number format
            if offset == 4:
                c.number_format = "0.0%"
            elif offset == 3:
                c.number_format = "@"   # text — Fav/Unfav label
            else:
                c.number_format = fmt

# ── Management Commentary ─────────────────────────────────────────────────────
ws.row_dimensions[16].height = 8
ws.row_dimensions[17].height = 18
c = ws.cell(row=17, column=1, value="MANAGEMENT COMMENTARY — Q1 FY2025")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A17:X17")

ws.row_dimensions[18].height = 14
c = ws.cell(row=18, column=1,
            value="2 paragraphs: (1) Revenue & Gross Profit drivers  "
                  "(2) OpEx & EBITDA drivers. Name specific months and AED amounts.")
c.fill = fill(MID_BLUE); c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A18:X18")

for r, para in [(19, "Paragraph 1 — Revenue & Gross Profit:"),
                (20, "Paragraph 2 — OpEx & EBITDA:")]:
    ws.row_dimensions[r].height = 50
    c = ws.cell(row=r, column=1, value=para)
    c.font = Font(color="7F7F7F", size=9, italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(f"A{r}:X{r}")

ws.freeze_panes = "B7"

wb.save(FILE)
print("Saved. Variance Analysis tab rebuilt with Fav/Unfav column.")
print("\nColumn layout per month block (e.g. January = cols B-F):")
print("  B = Budget    C = Actual    D = $ Variance    E = Fav/Unfav    F = % Variance")
print("  February: H-L    March: N-R    Q1 Total: T-X")
