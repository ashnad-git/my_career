"""
Zara & Co. FP&A Model — Workbook Generator
Phase 2B Project 1
Creates 08_fpa_model.xlsx with all 9 tabs structured and Assumptions tab populated.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "DCE6F1"   # hardcoded input cells
LIGHT_GREEN = "E2EFDA"   # formula cells (not used in Assumptions)
HEADER_BG   = "1F3864"
HEADER_FG   = "FFFFFF"
SECTION_BG  = "D6E4F0"
LABEL_BG    = "F2F2F2"
WHITE       = "FFFFFF"
NOTE_FG     = "7F7F7F"

# ── Style helpers ────────────────────────────────────────────────────────────
def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")

def font(bold=False, colour="000000", size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")

def align(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(bottom=s)

def set_cell(ws, row, col, value, bg=None, bold=False, colour="000000",
             size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, colour=colour, size=size, italic=italic)
    c.alignment = align(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c

def section_header(ws, row, col, text, end_col=None):
    c = set_cell(ws, row, col, text, bg=HEADER_BG, bold=True,
                 colour=HEADER_FG, size=9)
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    return c

def col_header(ws, row, col, text):
    return set_cell(ws, row, col, text, bg=SECTION_BG, bold=True,
                    colour=DARK_BLUE, size=9, h_align="center")

def input_cell(ws, row, col, value, number_format=None, h_align="center"):
    c = set_cell(ws, row, col, value, bg=LIGHT_BLUE, bold=False,
                 colour="1F497D", size=10, h_align=h_align,
                 number_format=number_format)
    return c

def label_cell(ws, row, col, text):
    return set_cell(ws, row, col, text, bg=LABEL_BG, bold=False,
                    colour="404040", size=10)

def note_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(color=NOTE_FG, size=9, italic=True, name="Calibri")
    c.alignment = align(horizontal="left", vertical="center")
    return c

# ── Workbook setup ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

TAB_DEFS = [
    ("Assumptions",       "4472C4"),   # blue
    ("Revenue Build",     "70AD47"),   # green
    ("COGS Build",        "FFC000"),   # yellow
    ("OpEx Detail",       "ED7D31"),   # orange
    ("P&L Summary",       "203864"),   # dark blue (light text)
    ("Variance Analysis", "FF0000"),   # red
    ("Rolling Forecast",  "7030A0"),   # purple
    ("Scenario Analysis", "A5A5A5"),   # grey
    ("Dashboard",         "00B0F0"),   # teal
]

# Create tabs in order
ws_assumptions = wb.active
ws_assumptions.title = "Assumptions"
ws_assumptions.sheet_properties.tabColor = "4472C4"

for name, colour in TAB_DEFS[1:]:
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = colour

# ── ASSUMPTIONS TAB ──────────────────────────────────────────────────────────
ws = ws_assumptions

# Column widths
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 22

# Row 1 — workbook title
ws.row_dimensions[1].height = 28
c = ws.cell(row=1, column=1, value="ZARA & CO. — FP&A MODEL FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=HEADER_FG, size=14, name="Calibri")
c.alignment = align(horizontal="left", vertical="center")
ws.merge_cells("A1:E1")

# Row 2 — sub-header
ws.row_dimensions[2].height = 16
c = ws.cell(row=2, column=1,
            value="Assumptions Tab — All inputs live here. Never hardcode numbers in other tabs.")
c.fill = fill(MID_BLUE)
c.font = Font(color=HEADER_FG, size=9, italic=True, name="Calibri")
c.alignment = align(horizontal="left", vertical="center")
ws.merge_cells("A2:E2")

# Row 3 — colour key legend
ws.row_dimensions[3].height = 16
set_cell(ws, 3, 1, "Colour key:", bold=True, size=9, colour=DARK_BLUE)
inp = ws.cell(row=3, column=2, value="  Blue = Hardcoded input  ")
inp.fill = fill(LIGHT_BLUE)
inp.font = Font(color="1F497D", size=9, name="Calibri")
inp.alignment = align(horizontal="center", vertical="center")

frm = ws.cell(row=3, column=3, value="  Green = Formula  ")
frm.fill = fill(LIGHT_GREEN)
frm.font = Font(color="375623", size=9, name="Calibri")
frm.alignment = align(horizontal="center", vertical="center")

set_cell(ws, 3, 4, "(Black = formula in other tabs)", italic=True,
         colour=NOTE_FG, size=9)

ws.row_dimensions[4].height = 8  # spacer

# ── SECTION A: Model Parameters ──────────────────────────────────────────────
ws.row_dimensions[5].height = 18
section_header(ws, 5, 1, "A. MODEL PARAMETERS", end_col=5)

ws.row_dimensions[6].height = 18
label_cell(ws, 6, 1, "Fiscal Year")
input_cell(ws, 6, 2, 2025, number_format="0")
label_cell(ws, 6, 3, "Currency")
input_cell(ws, 6, 4, "AED")

ws.row_dimensions[7].height = 18
label_cell(ws, 7, 1, "Company Name")
c = ws.cell(row=7, column=2, value="Zara & Co.")
c.fill = fill(LIGHT_BLUE)
c.font = Font(color="1F497D", size=10, name="Calibri")
c.alignment = align(horizontal="center")
ws.merge_cells("B7:D7")

ws.row_dimensions[8].height = 8  # spacer

# ── SECTION B: Outlets & Categories ─────────────────────────────────────────
ws.row_dimensions[9].height = 18
section_header(ws, 9, 1, "B. OUTLETS & PRODUCT CATEGORIES", end_col=5)

ws.row_dimensions[10].height = 18
label_cell(ws, 10, 1, "Outlet 1")
input_cell(ws, 10, 2, "Dubai Mall")
label_cell(ws, 10, 3, "Category 1")
input_cell(ws, 10, 4, "Women's Wear")

ws.row_dimensions[11].height = 18
label_cell(ws, 11, 1, "Outlet 2")
input_cell(ws, 11, 2, "Marina Mall")
label_cell(ws, 11, 3, "Category 2")
input_cell(ws, 11, 4, "Men's Wear")

ws.row_dimensions[12].height = 18
label_cell(ws, 12, 1, "Outlet 3")
input_cell(ws, 12, 2, "Online")
label_cell(ws, 12, 3, "Category 3")
input_cell(ws, 12, 4, "Kids' Wear")

ws.row_dimensions[13].height = 18
label_cell(ws, 13, 3, "Category 4")
input_cell(ws, 13, 4, "Accessories")

ws.row_dimensions[14].height = 8  # spacer

# ── SECTION C: Seasonality Multipliers ───────────────────────────────────────
ws.row_dimensions[15].height = 18
section_header(ws, 15, 1, "C. SEASONALITY MULTIPLIERS  (1.00 = base month)", end_col=5)

ws.row_dimensions[16].height = 16
col_header(ws, 16, 1, "Month")
col_header(ws, 16, 2, "Multiplier")
col_header(ws, 16, 3, "Note")

MONTHS = [
    ("January",   1.05, "Post-holiday shopping"),
    ("February",  1.00, ""),
    ("March",     1.20, "Ramadan — higher footfall & gifting"),
    ("April",     0.90, "Post-Ramadan dip"),
    ("May",       1.00, ""),
    ("June",      0.85, "Summer begins"),
    ("July",      0.80, "Peak summer — expat travel"),
    ("August",    0.85, "Summer"),
    ("September", 1.05, "Back to school / work"),
    ("October",   1.05, ""),
    ("November",  1.10, "White Friday / pre-holiday"),
    ("December",  1.15, "Holiday season + NYE"),
]

for i, (month, mult, note) in enumerate(MONTHS):
    r = 17 + i
    ws.row_dimensions[r].height = 17
    label_cell(ws, r, 1, month)
    input_cell(ws, r, 2, mult, number_format="0.00")
    if note:
        note_cell(ws, r, 3, note)

# Sum check row
r = 17 + len(MONTHS)
ws.row_dimensions[r].height = 17
set_cell(ws, r, 1, "Sum check (should ≈ 12)", bold=True, colour=DARK_BLUE,
         size=9, bg=SECTION_BG)
c = ws.cell(row=r, column=2, value=f"=SUM(B17:B{r-1})")
c.fill = fill(LIGHT_GREEN)
c.font = Font(color="375623", size=10, name="Calibri")
c.alignment = align(horizontal="center", vertical="center")
c.number_format = "0.00"

ws.row_dimensions[r + 1].height = 8  # spacer

# ── SECTION D: Base Monthly Units ────────────────────────────────────────────
row_d = r + 2
ws.row_dimensions[row_d].height = 18
section_header(ws, row_d, 1,
    "D. BASE MONTHLY UNITS  (per outlet, per category — before seasonality)", end_col=5)

row_d += 1
ws.row_dimensions[row_d].height = 16
col_header(ws, row_d, 1, "Category")
col_header(ws, row_d, 2, "Dubai Mall")
col_header(ws, row_d, 3, "Marina Mall")
col_header(ws, row_d, 4, "Online")

UNITS = [
    ("Women's Wear",  800,  600,  1200),
    ("Men's Wear",    500,  380,   700),
    ("Kids' Wear",    300,  250,   500),
    ("Accessories",   600,  450,   900),
]

row_d += 1
for cat, dm, mm, ol in UNITS:
    ws.row_dimensions[row_d].height = 17
    label_cell(ws, row_d, 1, cat)
    input_cell(ws, row_d, 2, dm, number_format="#,##0")
    input_cell(ws, row_d, 3, mm, number_format="#,##0")
    input_cell(ws, row_d, 4, ol, number_format="#,##0")
    row_d += 1

ws.row_dimensions[row_d].height = 8  # spacer
row_d += 1

# ── SECTION E: Average Selling Price ─────────────────────────────────────────
ws.row_dimensions[row_d].height = 18
section_header(ws, row_d, 1, "E. AVERAGE SELLING PRICE (AED)", end_col=5)

row_d += 1
ws.row_dimensions[row_d].height = 16
col_header(ws, row_d, 1, "Category")
col_header(ws, row_d, 2, "ASP (AED)")
col_header(ws, row_d, 3, "Note")

PRICES = [
    ("Women's Wear",  380, "Premium positioning — Dubai Mall mix weighted"),
    ("Men's Wear",    300, ""),
    ("Kids' Wear",    160, ""),
    ("Accessories",   130, "Impulse-buy category — lower ASP, high volume"),
]

row_d += 1
for cat, asp, note in PRICES:
    ws.row_dimensions[row_d].height = 17
    label_cell(ws, row_d, 1, cat)
    input_cell(ws, row_d, 2, asp, number_format="#,##0")
    if note:
        note_cell(ws, row_d, 3, note)
    row_d += 1

ws.row_dimensions[row_d].height = 8  # spacer
row_d += 1

# ── SECTION F: COGS % (Week 4 — placeholder) ─────────────────────────────────
ws.row_dimensions[row_d].height = 18
section_header(ws, row_d, 1, "F. COGS % BY CATEGORY  (Week 4 — fill in when building COGS tab)", end_col=5)

row_d += 1
col_header(ws, row_d, 1, "Category")
col_header(ws, row_d, 2, "COGS %")
row_d += 1

for cat in ["Women's Wear", "Men's Wear", "Kids' Wear", "Accessories"]:
    ws.row_dimensions[row_d].height = 17
    label_cell(ws, row_d, 1, cat)
    c = ws.cell(row=row_d, column=2, value="")
    c.fill = fill(LIGHT_BLUE)
    c.number_format = "0%"
    row_d += 1

ws.row_dimensions[row_d].height = 8
row_d += 1

# ── SECTION G: OpEx Assumptions (Week 4 — placeholder) ───────────────────────
ws.row_dimensions[row_d].height = 18
section_header(ws, row_d, 1, "G. OPEX ASSUMPTIONS (Monthly AED)  (Week 4)", end_col=5)

row_d += 1
col_header(ws, row_d, 1, "Item")
col_header(ws, row_d, 2, "Dubai Mall")
col_header(ws, row_d, 3, "Marina Mall")
col_header(ws, row_d, 4, "Online")
row_d += 1

for item in ["Rent (AED/month)", "Base Payroll (AED/month)", "Marketing % of Revenue"]:
    ws.row_dimensions[row_d].height = 17
    label_cell(ws, row_d, 1, item)
    for col in [2, 3, 4]:
        c = ws.cell(row=row_d, column=col, value="")
        c.fill = fill(LIGHT_BLUE)
    row_d += 1

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ── REVENUE BUILD TAB — skeleton only ────────────────────────────────────────
ws_rev = wb["Revenue Build"]

ws_rev.column_dimensions["A"].width = 20
ws_rev.column_dimensions["B"].width = 14
for col in range(3, 16):
    ws_rev.column_dimensions[get_column_letter(col)].width = 12

# Title
ws_rev.row_dimensions[1].height = 24
c = ws_rev.cell(row=1, column=1, value="ZARA & CO. — REVENUE BUILD FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=HEADER_FG, size=13, name="Calibri")
c.alignment = align(horizontal="left", vertical="center")
ws_rev.merge_cells("A1:O1")

ws_rev.row_dimensions[2].height = 14
c = ws_rev.cell(row=2, column=1,
    value="Formulas only — all inputs linked from Assumptions tab. Nothing hardcoded here.")
c.fill = fill(MID_BLUE)
c.font = Font(color=HEADER_FG, size=9, italic=True, name="Calibri")
c.alignment = align(horizontal="left", vertical="center")
ws_rev.merge_cells("A2:O2")

ws_rev.row_dimensions[3].height = 8

# Month headers row
ws_rev.row_dimensions[4].height = 16
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec","TOTAL"]
set_cell(ws_rev, 4, 1, "Outlet / Category", bg=SECTION_BG, bold=True,
         colour=DARK_BLUE, size=9, h_align="left")
set_cell(ws_rev, 4, 2, "Metric", bg=SECTION_BG, bold=True,
         colour=DARK_BLUE, size=9, h_align="center")
for i, m in enumerate(MONTH_ABBR):
    col_header(ws_rev, 4, 3 + i, m)

# Placeholder rows per outlet
OUTLET_NAMES = ["Dubai Mall", "Marina Mall", "Online"]
CAT_NAMES    = ["Women's Wear", "Men's Wear", "Kids' Wear", "Accessories"]
METRICS      = ["Units", "Revenue (AED)"]

row = 5
for outlet in OUTLET_NAMES:
    ws_rev.row_dimensions[row].height = 16
    c = ws_rev.cell(row=row, column=1, value=outlet)
    c.fill = fill(SECTION_BG)
    c.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
    ws_rev.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=15
    )
    row += 1

    for cat in CAT_NAMES:
        for metric in METRICS:
            ws_rev.row_dimensions[row].height = 16
            label_cell(ws_rev, row, 1, cat if metric == METRICS[0] else "")
            set_cell(ws_rev, row, 2, metric, colour="606060", size=9,
                     h_align="center")
            for col in range(3, 16):
                c = ws_rev.cell(row=row, column=col, value="")
                c.fill = fill(LIGHT_GREEN)
                c.font = Font(color="375623", size=10, name="Calibri")
                c.alignment = align(horizontal="center")
                if "Revenue" in metric:
                    c.number_format = "#,##0"
                else:
                    c.number_format = "#,##0"
            row += 1

    # Outlet subtotal row
    ws_rev.row_dimensions[row].height = 16
    set_cell(ws_rev, row, 1, f"{outlet} — Total Revenue", bold=True,
             colour=DARK_BLUE, bg=LIGHT_BLUE, size=10)
    set_cell(ws_rev, row, 2, "AED", bold=True, colour=DARK_BLUE,
             bg=LIGHT_BLUE, size=9, h_align="center")
    for col in range(3, 16):
        c = ws_rev.cell(row=row, column=col, value="")
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(bold=True, color="375623", size=10, name="Calibri")
        c.alignment = align(horizontal="center")
        c.number_format = "#,##0"
    row += 1
    row += 1  # spacer

# Grand total row
ws_rev.row_dimensions[row].height = 18
set_cell(ws_rev, row, 1, "TOTAL REVENUE — ALL OUTLETS", bold=True,
         colour=HEADER_FG, bg=HEADER_BG, size=10)
set_cell(ws_rev, row, 2, "AED", bold=True, colour=HEADER_FG,
         bg=HEADER_BG, size=9, h_align="center")
for col in range(3, 16):
    c = ws_rev.cell(row=row, column=col, value="")
    c.fill = fill(HEADER_BG)
    c.font = Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
    c.alignment = align(horizontal="center")
    c.number_format = "#,##0"

ws_rev.freeze_panes = "C5"

# ── Placeholder for remaining 7 tabs ─────────────────────────────────────────
PLACEHOLDER_TABS = [
    ("COGS Build",        "COGS Build — Week 4"),
    ("OpEx Detail",       "OpEx Detail — Week 4"),
    ("P&L Summary",       "P&L Summary — Week 4"),
    ("Variance Analysis", "Variance Analysis — Week 5"),
    ("Rolling Forecast",  "Rolling Forecast — Week 5"),
    ("Scenario Analysis", "Scenario Analysis — Week 6"),
    ("Dashboard",         "Dashboard — Week 6"),
]

for tab_name, placeholder_text in PLACEHOLDER_TABS:
    ws_t = wb[tab_name]
    ws_t.row_dimensions[1].height = 24
    c = ws_t.cell(row=1, column=1, value=f"ZARA & CO. — {placeholder_text.upper()}")
    c.fill = fill(DARK_BLUE)
    c.font = Font(bold=True, color=HEADER_FG, size=13, name="Calibri")
    c.alignment = align(horizontal="left", vertical="center")
    ws_t.merge_cells("A1:M1")
    ws_t.row_dimensions[2].height = 14
    c2 = ws_t.cell(row=2, column=1, value="Coming in " + placeholder_text.split("—")[1].strip())
    c2.fill = fill(MID_BLUE)
    c2.font = Font(color=HEADER_FG, size=9, italic=True, name="Calibri")
    c2.alignment = align(horizontal="left", vertical="center")
    ws_t.merge_cells("A2:M2")
    ws_t.column_dimensions["A"].width = 30

# ── Save ─────────────────────────────────────────────────────────────────────
OUTPUT = "/Users/ashnad/my_career/learning/phase2b_project1/08_fpa_model.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
