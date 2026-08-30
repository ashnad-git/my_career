"""
Adds three remaining tabs to 08_fpa_model.xlsx:
  1. Rolling Forecast  — Q1 actuals locked, Apr-Dec formula cells
  2. Scenario Analysis — Base / Bull / Bear with P&L output
  3. Dashboard         — KPI tiles + Q1 scorecard + FY outlook

Run: python3 learning/phase2b_project1/add_remaining_tabs.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILE = "learning/phase2b_project1/08_fpa_model.xlsx"

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_GREEN = "E2EFDA"
SECTION_BG  = "D6E4F0"
LABEL_BG    = "F2F2F2"
WHITE       = "FFFFFF"
ACTUAL_BG   = "FFF2CC"
TOTAL_BG    = "203864"
LIGHT_BLUE  = "DCE6F1"
BEAR_BG     = "FFE0E0"
TEAL_BG     = "DEEBF7"


def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def cell(ws, row, col, value=None, bg=None, bold=False, colour="000000",
         size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


def green_cell(ws, row, col, fmt="#,##0", bold=False):
    c = ws.cell(row=row, column=col, value=None)
    c.fill = fill(LIGHT_GREEN)
    c.font = Font(color="375623", size=10, name="Calibri", bold=bold)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = fmt
    return c


wb = openpyxl.load_workbook(FILE)

# ════════════════════════════════════════════════════════════════════════════
# TAB 1: ROLLING FORECAST
# ════════════════════════════════════════════════════════════════════════════
if "Rolling Forecast" in wb.sheetnames:
    del wb["Rolling Forecast"]
ws_rf = wb.create_sheet("Rolling Forecast", 6)
ws_rf.sheet_properties.tabColor = "7030A0"

ws_rf.column_dimensions["A"].width = 24
for letter in ["B", "C", "D"]:
    ws_rf.column_dimensions[letter].width = 13
for i in range(5, 14):
    ws_rf.column_dimensions[get_column_letter(i)].width = 13
ws_rf.column_dimensions["N"].width = 14

# Row 1: Title
ws_rf.row_dimensions[1].height = 24
c = ws_rf.cell(row=1, column=1, value="ZARA & CO. — ROLLING FORECAST FY2025")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_rf.merge_cells("A1:N1")

# Row 2: Sub-header
ws_rf.row_dimensions[2].height = 28
c = ws_rf.cell(row=2, column=1,
    value="Q1 actuals locked (Jan-Mar hardcoded). Apr-Dec = link from Revenue Build / COGS Build / OpEx Detail. "
          "FY Total = SUM of all 12 months. GP, EBITDA, margins = derived formulas.")
c.fill = fill(MID_BLUE); c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_rf.merge_cells("A2:N2")

# Row 3: Colour key
ws_rf.row_dimensions[3].height = 14
cell(ws_rf, 3, 1, "Colour key:", bold=True, size=9, colour=DARK_BLUE)
for col_n, label, bg, fg in [
    (2, "  Actual — locked  ", ACTUAL_BG, "7F6000"),
    (3, "  Forecast (formula)  ", LIGHT_GREEN, "375623"),
    (5, "  FY Total  ", TOTAL_BG, WHITE),
]:
    c = ws_rf.cell(row=3, column=col_n, value=label)
    c.fill = fill(bg); c.font = Font(color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

# Row 4: Spacer
ws_rf.row_dimensions[4].height = 8

# Row 5: Column headers
ws_rf.row_dimensions[5].height = 20
cell(ws_rf, 5, 1, "Line Item", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9)
for i, label in enumerate(["Jan Actual", "Feb Actual", "Mar Actual"]):
    c = ws_rf.cell(row=5, column=2+i, value=label)
    c.fill = fill(ACTUAL_BG); c.font = Font(bold=True, color="7F6000", size=8, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
for i, label in enumerate(["Apr Fcst","May Fcst","Jun Fcst","Jul Fcst","Aug Fcst",
                             "Sep Fcst","Oct Fcst","Nov Fcst","Dec Fcst"]):
    c = ws_rf.cell(row=5, column=5+i, value=label)
    c.fill = fill(SECTION_BG); c.font = Font(bold=True, color=DARK_BLUE, size=8, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
c = ws_rf.cell(row=5, column=14, value="FY2025 Total")
c.fill = fill(TOTAL_BG); c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")

# Row 6: Instruction note
ws_rf.row_dimensions[6].height = 20
c = ws_rf.cell(row=6, column=5,
    value="Apr-Dec: link from Revenue Build row 39 / COGS Build row 26 / OpEx Detail row 23  |  "
          "Col N: =SUM(B7:M7) for each row")
c.font = Font(color="7F7F7F", size=8, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_rf.merge_cells("E6:N6")

# P&L rows
PL_ROWS_RF = [
    (7,  "Revenue",          "#,##0", False),
    (8,  "Less: COGS",       "#,##0", False),
    (9,  "Gross Profit",     "#,##0", True),
    (10, "Gross Margin %",   "0.0%",  True),
    (12, "Less: OpEx",       "#,##0", False),
    (13, "EBITDA",           "#,##0", True),
    (14, "EBITDA Margin %",  "0.0%",  True),
]
ws_rf.row_dimensions[11].height = 6

ACTUALS_RF = {
    7:  (1_900_000, 1_860_000, 2_280_000),
    8:  (  858_000,   840_000, 1_040_000),
    12: (  685_000,   658_000,   710_000),
}

for tab_row, label, fmt, is_derived in PL_ROWS_RF:
    ws_rf.row_dimensions[tab_row].height = 17
    bold_label = is_derived or label in ["Revenue", "EBITDA"]

    cell(ws_rf, tab_row, 1, label,
         bg=SECTION_BG if is_derived else LABEL_BG,
         bold=bold_label,
         colour=DARK_BLUE if is_derived else "404040")

    # Jan/Feb/Mar actuals — hardcoded for source rows, green for derived
    for col_i, act_idx in [(2, 0), (3, 1), (4, 2)]:
        if tab_row in ACTUALS_RF:
            c = ws_rf.cell(row=tab_row, column=col_i, value=ACTUALS_RF[tab_row][act_idx])
            c.fill = fill(ACTUAL_BG)
            c.font = Font(color="7F6000", size=10, name="Calibri", bold=bold_label)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = fmt
        else:
            green_cell(ws_rf, tab_row, col_i, fmt=fmt, bold=bold_label)

    # Apr-Dec forecast — green formula cells
    for col_i in range(5, 14):
        green_cell(ws_rf, tab_row, col_i, fmt=fmt, bold=bold_label)

    # FY Total — dark blue formula cell
    c = ws_rf.cell(row=tab_row, column=14, value=None)
    c.fill = fill(TOTAL_BG)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = fmt

ws_rf.freeze_panes = "B7"


# ════════════════════════════════════════════════════════════════════════════
# TAB 2: SCENARIO ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
if "Scenario Analysis" in wb.sheetnames:
    del wb["Scenario Analysis"]
ws_sc = wb.create_sheet("Scenario Analysis", 7)
ws_sc.sheet_properties.tabColor = "ED7D31"

ws_sc.column_dimensions["A"].width = 30
ws_sc.column_dimensions["B"].width = 18  # Base
ws_sc.column_dimensions["C"].width = 18  # Bull
ws_sc.column_dimensions["D"].width = 18  # Bear
ws_sc.column_dimensions["E"].width = 3   # spacer
ws_sc.column_dimensions["F"].width = 18  # Bull vs Base
ws_sc.column_dimensions["G"].width = 18  # Bear vs Base

# Row 1: Title
ws_sc.row_dimensions[1].height = 24
c = ws_sc.cell(row=1, column=1, value="ZARA & CO. — SCENARIO ANALYSIS FY2025")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_sc.merge_cells("A1:G1")

# Row 2: Sub-header
ws_sc.row_dimensions[2].height = 28
c = ws_sc.cell(row=2, column=1,
    value="Base = budget assumptions linked from model. Bull = optimistic inputs. Bear = conservative inputs. "
          "P&L output in Section 2 recalculates automatically when you change Section 1 inputs.")
c.fill = fill(MID_BLUE); c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_sc.merge_cells("A2:G2")

# Row 3: Spacer
ws_sc.row_dimensions[3].height = 8

# ── SECTION 1: ASSUMPTIONS ──────────────────────────────────────────────────
ws_sc.row_dimensions[4].height = 18
c = ws_sc.cell(row=4, column=1, value="SECTION 1 — ASSUMPTIONS")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_sc.merge_cells("A4:G4")

ws_sc.row_dimensions[5].height = 16
for col_n, label, bg, fg in [
    (1, "Assumption",    SECTION_BG, DARK_BLUE),
    (2, "Base Case",     LIGHT_BLUE, "1F497D"),
    (3, "Bull Case",     LIGHT_GREEN, "375623"),
    (4, "Bear Case",     BEAR_BG, "9C0006"),
]:
    c = ws_sc.cell(row=5, column=col_n, value=label)
    c.fill = fill(bg); c.font = Font(bold=True, color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center" if col_n > 1 else "left", vertical="center")

ASSUMPTIONS = [
    (6,  "Full-Year Revenue (AED)",      "#,##0", "Link from P&L Summary col O, row 6"),
    (7,  "Revenue Growth % vs Budget",   "0.0%",  "Base = 0%. Bull/Bear: type the growth adjustment."),
    (8,  "Blended COGS % of Revenue",    "0.0%",  "Base: derive from model. Bull = lower %, Bear = higher %."),
    (9,  "Full-Year OpEx (AED)",         "#,##0", "Link from P&L Summary col O, row 11"),
    (10, "OpEx Growth % vs Budget",      "0.0%",  "Base = 0%. Bull = cost savings. Bear = overspend."),
]

for row, label, fmt, hint in ASSUMPTIONS:
    ws_sc.row_dimensions[row].height = 17
    cell(ws_sc, row, 1, label, bg=LABEL_BG, colour="404040")
    # Base: light blue
    c = ws_sc.cell(row=row, column=2, value=None)
    c.fill = fill(LIGHT_BLUE); c.font = Font(color="1F497D", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = fmt
    # Bull: green
    c = ws_sc.cell(row=row, column=3, value=None)
    c.fill = fill(LIGHT_GREEN); c.font = Font(color="375623", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = fmt
    # Bear: red
    c = ws_sc.cell(row=row, column=4, value=None)
    c.fill = fill(BEAR_BG); c.font = Font(color="9C0006", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = fmt
    # Hint text
    c = ws_sc.cell(row=row, column=6, value=hint)
    c.font = Font(color="7F7F7F", size=8, italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_sc.merge_cells(f"F{row}:G{row}")

# Row 11: Spacer
ws_sc.row_dimensions[11].height = 8

# ── SECTION 2: P&L OUTPUT ───────────────────────────────────────────────────
ws_sc.row_dimensions[12].height = 18
c = ws_sc.cell(row=12, column=1, value="SECTION 2 — P&L OUTPUT (formula-driven from Section 1)")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_sc.merge_cells("A12:G12")

ws_sc.row_dimensions[13].height = 16
for col_n, label, bg, fg in [
    (1, "Line Item",        SECTION_BG, DARK_BLUE),
    (2, "Base Case",        LIGHT_BLUE, "1F497D"),
    (3, "Bull Case",        LIGHT_GREEN, "375623"),
    (4, "Bear Case",        BEAR_BG, "9C0006"),
    (6, "Bull vs Base ($)", LIGHT_GREEN, "375623"),
    (7, "Bear vs Base ($)", BEAR_BG, "9C0006"),
]:
    c = ws_sc.cell(row=13, column=col_n, value=label)
    c.fill = fill(bg); c.font = Font(bold=True, color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center" if col_n > 1 else "left", vertical="center")

SC_PL_ROWS = [
    (14, "Revenue",          "#,##0", False),
    (15, "Less: COGS",       "#,##0", False),
    (16, "Gross Profit",     "#,##0", True),
    (17, "Gross Margin %",   "0.0%",  True),
    (19, "Less: OpEx",       "#,##0", False),
    (20, "EBITDA",           "#,##0", True),
    (21, "EBITDA Margin %",  "0.0%",  True),
]
ws_sc.row_dimensions[18].height = 6

for tab_row, label, fmt, is_derived in SC_PL_ROWS:
    ws_sc.row_dimensions[tab_row].height = 17
    bold_label = is_derived or label in ["Revenue", "EBITDA"]

    cell(ws_sc, tab_row, 1, label,
         bg=SECTION_BG if is_derived else LABEL_BG,
         bold=bold_label,
         colour=DARK_BLUE if is_derived else "404040")

    for col_n, bg, fg in [(2, LIGHT_BLUE, "1F497D"), (3, LIGHT_GREEN, "375623"), (4, BEAR_BG, "9C0006")]:
        c = ws_sc.cell(row=tab_row, column=col_n, value=None)
        c.fill = fill(bg); c.font = Font(color=fg, size=10, name="Calibri", bold=bold_label)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = fmt

    # Bull vs Base / Bear vs Base ($ impact — only for AED rows)
    if fmt != "0.0%":
        for col_n, bg, fg in [(6, LIGHT_GREEN, "375623"), (7, BEAR_BG, "9C0006")]:
            c = ws_sc.cell(row=tab_row, column=col_n, value=None)
            c.fill = fill(bg); c.font = Font(color=fg, size=10, name="Calibri", bold=bold_label)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = "#,##0"

ws_sc.freeze_panes = "B14"


# ════════════════════════════════════════════════════════════════════════════
# TAB 3: DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_db = wb.create_sheet("Dashboard", 8)
ws_db.sheet_properties.tabColor = "00B050"

ws_db.column_dimensions["A"].width = 24
ws_db.column_dimensions["B"].width = 17  # Budget
ws_db.column_dimensions["C"].width = 17  # Actual / Forecast
ws_db.column_dimensions["D"].width = 14  # $ Variance
ws_db.column_dimensions["E"].width = 11  # Fav/Unfav
ws_db.column_dimensions["F"].width = 11  # % Variance

# Row 1: Title
ws_db.row_dimensions[1].height = 24
c = ws_db.cell(row=1, column=1, value="ZARA & CO. — EXECUTIVE DASHBOARD FY2025")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_db.merge_cells("A1:F1")

# Row 2: Sub-header
ws_db.row_dimensions[2].height = 14
c = ws_db.cell(row=2, column=1,
    value="All cells are formulas — links from P&L Summary (budget), Variance Analysis (Q1 actuals), Rolling Forecast (FY).")
c.fill = fill(MID_BLUE); c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_db.merge_cells("A2:F2")

# Row 3: Spacer
ws_db.row_dimensions[3].height = 8

# ── SECTION 1: KPI TILES ────────────────────────────────────────────────────
ws_db.row_dimensions[4].height = 18
c = ws_db.cell(row=4, column=1, value="SECTION 1 — Q1 KEY METRICS")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_db.merge_cells("A4:F4")

ws_db.row_dimensions[5].height = 14   # KPI label row
ws_db.row_dimensions[6].height = 32   # KPI value row (tall)
ws_db.row_dimensions[7].height = 14   # source hint row

KPI_DATA = [
    (2, "Q1 Revenue — Actual",   "#,##0", "='Variance Analysis'!U7"),
    (3, "Q1 EBITDA — Actual",    "#,##0", "='Variance Analysis'!U13"),
    (4, "Q1 Gross Margin %",     "0.0%",  "='Variance Analysis'!U10"),
    (5, "Q1 EBITDA Margin %",    "0.0%",  "='Variance Analysis'!U14"),
]

for col_n, label, fmt, src in KPI_DATA:
    c = ws_db.cell(row=5, column=col_n, value=label)
    c.fill = fill(MID_BLUE); c.font = Font(bold=True, color=WHITE, size=8, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    c = ws_db.cell(row=6, column=col_n, value=None)
    c.fill = fill(LIGHT_GREEN)
    c.font = Font(bold=True, color="375623", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = fmt

    c = ws_db.cell(row=7, column=col_n, value=f"Source: {src}")
    c.fill = fill(MID_BLUE); c.font = Font(color=WHITE, size=7, italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

# Row 8: Spacer
ws_db.row_dimensions[8].height = 8

# ── SECTION 2: Q1 SCORECARD ─────────────────────────────────────────────────
ws_db.row_dimensions[9].height = 18
c = ws_db.cell(row=9, column=1, value="SECTION 2 — Q1 SCORECARD (Actual vs Budget)")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_db.merge_cells("A9:F9")

ws_db.row_dimensions[10].height = 14
for col_n, label, bg, fg in [
    (1, "Line Item",   SECTION_BG, DARK_BLUE),
    (2, "Q1 Budget",   LIGHT_BLUE, "1F497D"),
    (3, "Q1 Actual",   ACTUAL_BG, "7F6000"),
    (4, "$ Variance",  LIGHT_GREEN, "375623"),
    (5, "Fav/Unfav",   LIGHT_GREEN, "375623"),
    (6, "% Variance",  LIGHT_GREEN, "375623"),
]:
    c = ws_db.cell(row=10, column=col_n, value=label)
    c.fill = fill(bg); c.font = Font(bold=True, color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center" if col_n > 1 else "left", vertical="center")

Q1_ROWS = [
    (11, "Revenue",          "#,##0", False),
    (12, "Less: COGS",       "#,##0", False),
    (13, "Gross Profit",     "#,##0", True),
    (14, "Gross Margin %",   "0.0%",  True),
    (16, "Less: OpEx",       "#,##0", False),
    (17, "EBITDA",           "#,##0", True),
    (18, "EBITDA Margin %",  "0.0%",  True),
]
ws_db.row_dimensions[15].height = 6

for tab_row, label, fmt, is_derived in Q1_ROWS:
    ws_db.row_dimensions[tab_row].height = 17
    bold_label = is_derived or label in ["Revenue", "EBITDA"]

    cell(ws_db, tab_row, 1, label,
         bg=SECTION_BG if is_derived else LABEL_BG,
         bold=bold_label,
         colour=DARK_BLUE if is_derived else "404040")

    for col_n, bg, fg, num_fmt in [
        (2, LIGHT_BLUE,  "1F497D", fmt),
        (3, ACTUAL_BG,   "7F6000", fmt),
        (4, LIGHT_GREEN, "375623", fmt if fmt != "0.0%" else "#,##0"),
        (5, LIGHT_GREEN, "375623", "@"),
        (6, LIGHT_GREEN, "375623", "0.0%"),
    ]:
        c = ws_db.cell(row=tab_row, column=col_n, value=None)
        c.fill = fill(bg); c.font = Font(color=fg, size=10, name="Calibri", bold=bold_label)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = num_fmt

# Row 19: Spacer
ws_db.row_dimensions[19].height = 8

# ── SECTION 3: FY OUTLOOK ───────────────────────────────────────────────────
ws_db.row_dimensions[20].height = 18
c = ws_db.cell(row=20, column=1, value="SECTION 3 — FULL YEAR OUTLOOK (Budget vs Rolling Forecast)")
c.fill = fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_db.merge_cells("A20:F20")

ws_db.row_dimensions[21].height = 14
for col_n, label, bg, fg in [
    (1, "Line Item",          SECTION_BG, DARK_BLUE),
    (2, "FY2025 Budget",      LIGHT_BLUE, "1F497D"),
    (3, "Rolling Forecast",   LIGHT_GREEN, "375623"),
    (4, "$ Variance",         LIGHT_GREEN, "375623"),
    (5, "Fav/Unfav",          LIGHT_GREEN, "375623"),
    (6, "% Variance",         LIGHT_GREEN, "375623"),
]:
    c = ws_db.cell(row=21, column=col_n, value=label)
    c.fill = fill(bg); c.font = Font(bold=True, color=fg, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center" if col_n > 1 else "left", vertical="center")

FY_ROWS = [
    (22, "Revenue",          "#,##0", False),
    (23, "Less: COGS",       "#,##0", False),
    (24, "Gross Profit",     "#,##0", True),
    (25, "Gross Margin %",   "0.0%",  True),
    (27, "Less: OpEx",       "#,##0", False),
    (28, "EBITDA",           "#,##0", True),
    (29, "EBITDA Margin %",  "0.0%",  True),
]
ws_db.row_dimensions[26].height = 6

for tab_row, label, fmt, is_derived in FY_ROWS:
    ws_db.row_dimensions[tab_row].height = 17
    bold_label = is_derived or label in ["Revenue", "EBITDA"]

    cell(ws_db, tab_row, 1, label,
         bg=SECTION_BG if is_derived else LABEL_BG,
         bold=bold_label,
         colour=DARK_BLUE if is_derived else "404040")

    for col_n, bg, fg, num_fmt in [
        (2, LIGHT_BLUE,  "1F497D", fmt),
        (3, LIGHT_GREEN, "375623", fmt),
        (4, LIGHT_GREEN, "375623", fmt if fmt != "0.0%" else "#,##0"),
        (5, LIGHT_GREEN, "375623", "@"),
        (6, LIGHT_GREEN, "375623", "0.0%"),
    ]:
        c = ws_db.cell(row=tab_row, column=col_n, value=None)
        c.fill = fill(bg); c.font = Font(color=fg, size=10, name="Calibri", bold=bold_label)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.number_format = num_fmt

ws_db.freeze_panes = "B11"

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(FILE)
print("Saved. Three tabs added to 08_fpa_model.xlsx")
print()
print("ROLLING FORECAST (tab 7):")
print("  B/C/D = Jan/Feb/Mar actuals locked (yellow)")
print("  E-M   = Apr-Dec forecast (green formula cells)")
print("  N     = FY2025 Total (dark blue formula cell)")
print("  Source rows: Revenue Build row 39 | COGS Build row 26 | OpEx Detail row 23")
print()
print("SCENARIO ANALYSIS (tab 8):")
print("  Section 1 (rows 6-10): Assumptions — Base=blue, Bull=green, Bear=red")
print("  Section 2 (rows 14-21): P&L Output + Bull/Bear vs Base impact columns")
print()
print("DASHBOARD (tab 9):")
print("  Section 1 (rows 5-7): 4 KPI tiles from Variance Analysis Q1 block")
print("  Section 2 (rows 11-18): Q1 Scorecard — link from Variance Analysis Q1 (cols T-X)")
print("  Section 3 (rows 22-29): FY Outlook — Budget from P&L Summary, Forecast from Rolling Forecast")
