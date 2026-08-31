"""
Fills Assumptions Section G (OpEx inputs) and builds OpEx Detail tab template.
Run: python3 learning/phase2b_project1/add_opex_tab.py

Assumptions Section G:
  Row 54: Rent (AED/month)         — B=Dubai Mall, C=Marina Mall, D=Online
  Row 55: Base Payroll (AED/month) — B=Dubai Mall, C=Marina Mall, D=Online
  Row 56: Marketing % of Revenue   — B=Dubai Mall, C=Marina Mall, D=Online

Revenue Build outlet total rows (confirmed from actual file):
  Row 15: Dubai Mall Total Revenue
  Row 26: Marina Mall Total Revenue
  Row 37: Online Total Revenue

Column layout matches COGS Build — Jan=C, Feb=D, ... Dec=N, TOTAL=O
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILE = "/Users/ashnad/my_career/learning/phase2b_project1/08_fpa_model.xlsx"

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "DCE6F1"
LIGHT_GREEN = "E2EFDA"
SECTION_BG  = "D6E4F0"
LABEL_BG    = "F2F2F2"
WHITE       = "FFFFFF"


def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def set_cell(ws, row, col, value, bg=None, bold=False, colour="000000",
             size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


wb = openpyxl.load_workbook(FILE)

# ── Step 1: Fill Assumptions Section G ───────────────────────────────────────
ws_a = wb["Assumptions"]

OPEX_INPUTS = [
    # (row, DM_val, MM_val, OL_val, fmt)
    (54, 180_000, 120_000, 15_000,  "#,##0"),    # Rent
    (55, 150_000, 100_000, 45_000,  "#,##0"),    # Payroll
    (56,    0.03,    0.03,   0.02,  "0%"),       # Marketing %
]

OUTLET_LABELS = ["Dubai Mall", "Marina Mall", "Online"]
for row, dm, mm, ol, fmt in OPEX_INPUTS:
    label = ws_a.cell(row=row, column=1).value
    for col_idx, val in enumerate([dm, mm, ol], start=2):
        c = ws_a.cell(row=row, column=col_idx, value=val)
        c.fill = fill(LIGHT_BLUE)
        c.font = Font(color="1F497D", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
    print(f"  Assumptions row {row} ({label}): DM={dm}, MM={mm}, OL={ol}")

# ── Step 2: Recreate OpEx Detail tab ─────────────────────────────────────────
if "OpEx Detail" in wb.sheetnames:
    tab_idx = wb.sheetnames.index("OpEx Detail")
    del wb["OpEx Detail"]
else:
    tab_idx = 3

ws = wb.create_sheet("OpEx Detail", tab_idx)
ws.sheet_properties.tabColor = "ED7D31"

# Column widths — matches COGS Build (no col B metric column for OpEx)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 4   # narrow spacer, unused
for col in range(3, 16):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Row 1: Title
ws.row_dimensions[1].height = 24
c = ws.cell(row=1, column=1, value="ZARA & CO. — OPEX DETAIL FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:O1")

# Row 2: Sub-header
ws.row_dimensions[2].height = 14
c = ws.cell(row=2, column=1,
            value="Formulas only — fixed costs from Assumptions, marketing linked to Revenue Build.")
c.fill = fill(MID_BLUE)
c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A2:O2")

# Row 3: Spacer
ws.row_dimensions[3].height = 8

# Row 4: Column headers — col A = Item, cols C-O = Jan-Dec + TOTAL (col B skipped)
ws.row_dimensions[4].height = 16
set_cell(ws, 4, 1, "Item", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9)
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","TOTAL"]
for i, m in enumerate(MONTHS):
    set_cell(ws, 4, 3+i, m, bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9, h_align="center")

# Outlet data
ITEMS = ["Rent", "Payroll", "Marketing"]

OUTLETS = [
    {"name": "Dubai Mall",   "header_row": 5,  "data_rows": [6,7,8],   "total_row": 9,  "spacer": 10},
    {"name": "Marina Mall",  "header_row": 11, "data_rows": [12,13,14],"total_row": 15, "spacer": 16},
    {"name": "Online",       "header_row": 17, "data_rows": [18,19,20],"total_row": 21, "spacer": 22},
]
GRAND_TOTAL_ROW = 23

for outlet in OUTLETS:
    # Outlet header
    r = outlet["header_row"]
    ws.row_dimensions[r].height = 16
    c = ws.cell(row=r, column=1, value=outlet["name"])
    c.fill = fill(SECTION_BG)
    c.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)

    # Item rows
    for data_row, item in zip(outlet["data_rows"], ITEMS):
        ws.row_dimensions[data_row].height = 16
        set_cell(ws, data_row, 1, item, bg=LABEL_BG, colour="404040")

        # Monthly formula cells (cols C-N) — empty green placeholders
        for col_idx in range(3, 15):
            c = ws.cell(row=data_row, column=col_idx, value=None)
            c.fill = fill(LIGHT_GREEN)
            c.font = Font(color="375623", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"

        # Total column (O) — empty green placeholder
        c = ws.cell(row=data_row, column=15, value=None)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(bold=True, color="375623", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"

    # Outlet total row
    tr = outlet["total_row"]
    ws.row_dimensions[tr].height = 16
    set_cell(ws, tr, 1, f"{outlet['name']} — Total OpEx",
             bold=True, colour=DARK_BLUE, bg=LIGHT_BLUE)
    for col_idx in range(3, 16):
        c = ws.cell(row=tr, column=col_idx, value=None)
        c.fill = fill(LIGHT_BLUE)
        c.font = Font(bold=True, color="1F497D", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"

    # Spacer
    ws.row_dimensions[outlet["spacer"]].height = 8

# Grand total row
ws.row_dimensions[GRAND_TOTAL_ROW].height = 18
c = ws.cell(row=GRAND_TOTAL_ROW, column=1, value="TOTAL OPEX — ALL OUTLETS")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
for col_idx in range(3, 16):
    c = ws.cell(row=GRAND_TOTAL_ROW, column=col_idx, value=None)
    c.fill = fill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "#,##0"

ws.freeze_panes = "C5"

wb.save(FILE)
print("\nSaved. Open 08_fpa_model.xlsx — OpEx Detail tab ready for formulas.")
