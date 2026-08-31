"""
Fills Assumptions COGS% (Section F) and builds the full COGS Build tab.
Run: python3 learning/phase2b_project1/add_cogs_tab.py

Revenue Build row reference (from generator structure):
  Dubai Mall   — Women's 7, Men's 9, Kids' 11, Accessories 13
  Marina Mall  — Women's 18, Men's 20, Kids' 22, Accessories 24
  Online       — Women's 29, Men's 31, Kids' 33, Accessories 35
  (Revenue rows only — each category has a Units row above it)

Assumptions COGS% cells:
  B47 Women's Wear, B48 Men's Wear, B49 Kids' Wear, B50 Accessories
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter

FILE = "/Users/ashnad/my_career/learning/phase2b_project1/08_fpa_model.xlsx"

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "DCE6F1"
LIGHT_GREEN = "E2EFDA"
SECTION_BG = "D6E4F0"
LABEL_BG   = "F2F2F2"
WHITE      = "FFFFFF"


def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def set_cell(ws, row, col, value, bg=None, bold=False, colour="000000",
             size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


wb = openpyxl.load_workbook(FILE)

# ── Step 1: Fill COGS% in Assumptions Section F ───────────────────────────────
ws_a = wb["Assumptions"]

COGS_INPUTS = [
    (47, 0.45, "Women's Wear"),
    (48, 0.48, "Men's Wear"),
    (49, 0.50, "Kids' Wear"),
    (50, 0.35, "Accessories"),
]
for row, rate, cat in COGS_INPUTS:
    c = ws_a.cell(row=row, column=2, value=rate)
    c.fill = fill(LIGHT_BLUE)
    c.font = Font(color="1F497D", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0%"
    print(f"  Assumptions B{row}: {cat} COGS% = {rate:.0%}")

# ── Step 2: Recreate COGS Build tab cleanly ───────────────────────────────────
if "COGS Build" in wb.sheetnames:
    tab_idx = wb.sheetnames.index("COGS Build")
    del wb["COGS Build"]
else:
    tab_idx = 2

ws = wb.create_sheet("COGS Build", tab_idx)
ws.sheet_properties.tabColor = "FFC000"

# Column widths
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 10
for col in range(3, 16):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Row 1: Title
ws.row_dimensions[1].height = 24
c = ws.cell(row=1, column=1, value="ZARA & CO. — COGS BUILD FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:O1")

# Row 2: Sub-header
ws.row_dimensions[2].height = 14
c = ws.cell(row=2, column=1,
            value="Formulas only — COGS% pulled from Assumptions, amounts = Revenue × COGS%.")
c.fill = fill(MID_BLUE)
c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A2:O2")

# Row 3: Spacer
ws.row_dimensions[3].height = 8

# Row 4: Column headers
ws.row_dimensions[4].height = 16
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","TOTAL"]
set_cell(ws, 4, 1, "Outlet / Category", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9)
set_cell(ws, 4, 2, "COGS %", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9, h_align="center")
for i, m in enumerate(MONTHS):
    set_cell(ws, 4, 3+i, m, bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9, h_align="center")

# ── Outlet / category data ────────────────────────────────────────────────────
CATS = ["Women's Wear", "Men's Wear", "Kids' Wear", "Accessories"]
COGS_PCT_ROWS = [47, 48, 49, 50]   # Assumptions col B rows

OUTLETS = [
    {
        "name":       "Dubai Mall",
        "header_row": 5,
        "data_rows":  [6, 7, 8, 9],
        "total_row":  10,
        "spacer_row": 11,
        "rev_rows":   [7, 9, 11, 13],   # Revenue Build revenue rows per category
    },
    {
        "name":       "Marina Mall",
        "header_row": 12,
        "data_rows":  [13, 14, 15, 16],
        "total_row":  17,
        "spacer_row": 18,
        "rev_rows":   [18, 20, 22, 24],
    },
    {
        "name":       "Online",
        "header_row": 19,
        "data_rows":  [20, 21, 22, 23],
        "total_row":  24,
        "spacer_row": 25,
        "rev_rows":   [29, 31, 33, 35],
    },
]
GRAND_TOTAL_ROW = 26

for outlet in OUTLETS:
    # Outlet header
    r = outlet["header_row"]
    ws.row_dimensions[r].height = 16
    c = ws.cell(row=r, column=1, value=outlet["name"])
    c.fill = fill(SECTION_BG)
    c.font = Font(bold=True, color=DARK_BLUE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)

    # Category rows
    for data_row, cat, rev_row, pct_row in zip(
        outlet["data_rows"], CATS, outlet["rev_rows"], COGS_PCT_ROWS
    ):
        ws.row_dimensions[data_row].height = 16

        # Col A: label
        set_cell(ws, data_row, 1, cat, bg=LABEL_BG, colour="404040")

        # Col B: COGS% — empty placeholder, Ashnad links from Assumptions
        c = ws.cell(row=data_row, column=2, value=None)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(color="375623", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0%"

        # Cols C–N: monthly COGS — empty placeholders, Ashnad writes the formulas
        for col_idx in range(3, 15):
            c = ws.cell(row=data_row, column=col_idx, value=None)
            c.fill = fill(LIGHT_GREEN)
            c.font = Font(color="375623", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"

        # Col O: total — empty placeholder
        c = ws.cell(row=data_row, column=15, value=None)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(bold=True, color="375623", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"

    # Outlet total row
    tr = outlet["total_row"]
    first_d = outlet["data_rows"][0]
    last_d  = outlet["data_rows"][-1]
    ws.row_dimensions[tr].height = 16

    set_cell(ws, tr, 1, f"{outlet['name']} — Total COGS",
             bold=True, colour=DARK_BLUE, bg=LIGHT_BLUE)
    set_cell(ws, tr, 2, "", bg=LIGHT_BLUE)

    for col_idx in range(3, 16):
        c = ws.cell(row=tr, column=col_idx, value=None)
        c.fill = fill(LIGHT_BLUE)
        c.font = Font(bold=True, color="1F497D", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "#,##0"

    # Spacer
    ws.row_dimensions[outlet["spacer_row"]].height = 8

# Grand total row
ws.row_dimensions[GRAND_TOTAL_ROW].height = 18
c = ws.cell(row=GRAND_TOTAL_ROW, column=1, value="TOTAL COGS — ALL OUTLETS")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")

ws.cell(row=GRAND_TOTAL_ROW, column=2).fill = fill(DARK_BLUE)

outlet_total_rows = [o["total_row"] for o in OUTLETS]   # [10, 17, 24]

for col_idx in range(3, 16):
    c = ws.cell(row=GRAND_TOTAL_ROW, column=col_idx, value=None)
    c.fill = fill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "#,##0"

ws.freeze_panes = "C5"

wb.save(FILE)
print("\nSaved. Open 08_fpa_model.xlsx — check COGS Build tab and Assumptions Section F.")
