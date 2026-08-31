"""
Builds P&L Summary tab template.
Run: python3 learning/phase2b_project1/add_pl_summary_tab.py

Source rows (confirmed from actual file):
  Revenue Build  grand total: row 38 (cols C-N = Jan-Dec, col O = Annual)
  COGS Build     grand total: row 26
  OpEx Detail    grand total: row 23

Column layout:
  A = Label
  B = FY2024 Prior Year (hardcoded — pre-filled)
  C-N = Jan-Dec Budget (formula cells — Ashnad fills)
  O = FY2025 Annual Budget (formula cell — Ashnad fills)
  P = Jan Actual (hardcoded — pre-filled)
  Q = Feb Actual (hardcoded — pre-filled)
  R = Mar Actual (hardcoded — pre-filled)

P&L rows:
  6  = Revenue
  7  = Less: COGS
  8  = Gross Profit       (formula: Revenue - COGS)
  9  = Gross Margin %     (formula: Gross Profit / Revenue)
  10 = spacer
  11 = Less: OpEx
  12 = EBITDA             (formula: Gross Profit - OpEx)
  13 = EBITDA Margin %    (formula: EBITDA / Revenue)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILE = "/Users/ashnad/my_career/learning/phase2b_project1/08_fpa_model.xlsx"

DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "DCE6F1"
LIGHT_GREEN = "E2EFDA"
SECTION_BG  = "D6E4F0"
LABEL_BG    = "F2F2F2"
TOTAL_BG    = "203864"   # dark blue for annual column
WHITE       = "FFFFFF"
ACTUAL_BG   = "FFF2CC"  # yellow — actuals are different from budget


def fill(hex_colour):
    return PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")


def set_cell(ws, row, col, value, bg=None, bold=False, colour="000000",
             size=10, h_align="left", italic=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if number_format:
        c.number_format = number_format
    return c


wb = openpyxl.load_workbook(FILE)

if "P&L Summary" in wb.sheetnames:
    tab_idx = wb.sheetnames.index("P&L Summary")
    del wb["P&L Summary"]
else:
    tab_idx = 4

ws = wb.create_sheet("P&L Summary", tab_idx)
ws.sheet_properties.tabColor = "203864"

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 16   # FY2024
for col in range(3, 16):               # Jan-Dec + Annual (C-O)
    ws.column_dimensions[get_column_letter(col)].width = 13
ws.column_dimensions["P"].width = 13  # Jan Actual
ws.column_dimensions["Q"].width = 13  # Feb Actual
ws.column_dimensions["R"].width = 13  # Mar Actual

# ── Row 1: Title ──────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 24
c = ws.cell(row=1, column=1, value="ZARA & CO. — P&L SUMMARY FY2025")
c.fill = fill(DARK_BLUE)
c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A1:R1")

# ── Row 2: Sub-header ─────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 14
c = ws.cell(row=2, column=1,
            value="Budget linked from Revenue Build / COGS Build / OpEx Detail. Actuals hardcoded.")
c.fill = fill(MID_BLUE)
c.font = Font(color=WHITE, size=9, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells("A2:R2")

# ── Row 3: Colour key ─────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 14
set_cell(ws, 3, 1, "Colour key:", bold=True, size=9, colour=DARK_BLUE)
c = ws.cell(row=3, column=2, value="  FY2024 Prior Year  ")
c.fill = fill(LIGHT_BLUE); c.font = Font(color="1F497D", size=9, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")
c = ws.cell(row=3, column=3, value="  Budget (formula)  ")
c.fill = fill(LIGHT_GREEN); c.font = Font(color="375623", size=9, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")
c = ws.cell(row=3, column=4, value="  Actual (hardcoded)  ")
c.fill = fill(ACTUAL_BG); c.font = Font(color="7F6000", size=9, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 4: Spacer ─────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 8

# ── Row 5: Column headers ─────────────────────────────────────────────────────
ws.row_dimensions[5].height = 16
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
set_cell(ws, 5, 1, "Line Item", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9)
set_cell(ws, 5, 2, "FY2024 Actual", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=9, h_align="center")
for i, m in enumerate(MONTHS):
    set_cell(ws, 5, 3+i, f"{m} Budget", bg=SECTION_BG, bold=True, colour=DARK_BLUE, size=8, h_align="center")
set_cell(ws, 5, 15, "FY2025 Budget", bg=TOTAL_BG, bold=True, colour=WHITE, size=9, h_align="center")
set_cell(ws, 5, 16, "Jan Actual", bg=ACTUAL_BG, bold=True, colour="7F6000", size=9, h_align="center")
set_cell(ws, 5, 17, "Feb Actual", bg=ACTUAL_BG, bold=True, colour="7F6000", size=9, h_align="center")
set_cell(ws, 5, 18, "Mar Actual", bg=ACTUAL_BG, bold=True, colour="7F6000", size=9, h_align="center")

# ── P&L line items ────────────────────────────────────────────────────────────
# Prior Year hardcoded values (FY2024)
PRIOR_YEAR = {
    6:  19_800_000,   # Revenue
    7:   9_200_000,   # COGS
    11:  7_100_000,   # OpEx
}

# Actuals hardcoded (Jan/Feb/Mar — slight variance from budget for interesting analysis)
ACTUALS = {
    #        Jan         Feb         Mar
    6:  (1_900_000,  1_860_000,  2_280_000),   # Revenue (Jan/Feb miss, Mar Ramadan beat)
    7:  (  858_000,    840_000,  1_040_000),   # COGS
    11: (  685_000,    658_000,    710_000),   # OpEx (slight overspend on marketing)
}

ROWS = [
    (6,  "Revenue",          "#,##0",  False),
    (7,  "Less: COGS",       "#,##0",  False),
    (8,  "Gross Profit",     "#,##0",  True),
    (9,  "Gross Margin %",   "0.0%",   True),
    (11, "Less: OpEx",       "#,##0",  False),
    (12, "EBITDA",           "#,##0",  True),
    (13, "EBITDA Margin %",  "0.0%",   True),
]

for row, label, fmt, is_derived in ROWS:
    ws.row_dimensions[row].height = 17

    # Col A: label
    bold_label = is_derived or row in [6, 12]
    set_cell(ws, row, 1, label, bg=LABEL_BG if not is_derived else SECTION_BG,
             bold=bold_label, colour=DARK_BLUE if is_derived else "404040")

    # Col B: FY2024 prior year
    if row in PRIOR_YEAR:
        c = ws.cell(row=row, column=2, value=PRIOR_YEAR[row])
        c.fill = fill(LIGHT_BLUE)
        c.font = Font(color="1F497D", size=10, name="Calibri", bold=bold_label)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
    else:
        # Derived prior year rows (GP, margins, EBITDA) — empty green for Ashnad
        c = ws.cell(row=row, column=2, value=None)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(color="375623", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

    # Cols C-N: monthly budget — empty green formula cells
    for col_idx in range(3, 15):
        c = ws.cell(row=row, column=col_idx, value=None)
        c.fill = fill(LIGHT_GREEN)
        c.font = Font(color="375623", size=10, name="Calibri", bold=bold_label)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt

    # Col O: annual budget — empty green
    c = ws.cell(row=row, column=15, value=None)
    c.fill = fill(TOTAL_BG)
    c.font = Font(color=WHITE, size=10, name="Calibri", bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = fmt

    # Cols P-R: Jan/Feb/Mar actual — pre-filled if source row, green if derived
    for act_col, act_idx in [(16, 0), (17, 1), (18, 2)]:
        if row in ACTUALS:
            c = ws.cell(row=row, column=act_col, value=ACTUALS[row][act_idx])
            c.fill = fill(ACTUAL_BG)
            c.font = Font(color="7F6000", size=10, name="Calibri", bold=bold_label)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = fmt
        else:
            c = ws.cell(row=row, column=act_col, value=None)
            c.fill = fill(LIGHT_GREEN)
            c.font = Font(color="375623", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = fmt

# Row 10: spacer between GP and OpEx
ws.row_dimensions[10].height = 6

ws.freeze_panes = "C6"

wb.save(FILE)
print("Saved. Open 08_fpa_model.xlsx — P&L Summary tab ready.")
print("\nWhat Ashnad needs to fill:")
print("  Green cells = formulas to write")
print("  Blue = FY2024 prior year (pre-filled)")
print("  Yellow = Jan/Feb/Mar actuals (pre-filled)")
print("\nSource rows:")
print("  Revenue:  Revenue Build  row 38  (cols C-N=months, O=annual)")
print("  COGS:     COGS Build     row 26")
print("  OpEx:     OpEx Detail    row 23")
